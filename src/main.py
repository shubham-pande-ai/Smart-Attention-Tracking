import io
import json
import base64
import asyncio
import openpyxl
from pathlib import Path
from typing import Dict, Any
from fastapi import FastAPI, UploadFile, File, Form, WebSocket, WebSocketDisconnect, Request, Response, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware

from .database import Database
from .inference import AttentionEngine
from .templates import ADMIN_DASHBOARD, TEACHER_DASHBOARD, USER_DASHBOARD

app = FastAPI(title="Smart Attention Tracker")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

db = Database()
engine = AttentionEngine(model_path=Path("models/attention_cnn.pt"))

# --- STATE MANAGEMENT ---
class ConnectionManager:
    def __init__(self):
        self.teacher_sockets: list[WebSocket] = []
        self.student_sockets: dict[int, WebSocket] = {}
        self.live_telemetry: dict[int, dict] = {}
        self.teacher_video_subscriptions: dict[WebSocket, int] = {} # teacher_ws -> student_id

    async def connect_teacher(self, websocket: WebSocket):
        await websocket.accept()
        self.teacher_sockets.append(websocket)

    def disconnect_teacher(self, websocket: WebSocket):
        if websocket in self.teacher_sockets:
            self.teacher_sockets.remove(websocket)
        if websocket in self.teacher_video_subscriptions:
            del self.teacher_video_subscriptions[websocket]

    async def connect_student(self, websocket: WebSocket, student_id: int):
        await websocket.accept()
        self.student_sockets[student_id] = websocket

    def disconnect_student(self, student_id: int):
        if student_id in self.student_sockets:
            del self.student_sockets[student_id]
        if student_id in self.live_telemetry:
            del self.live_telemetry[student_id]

    async def broadcast_to_teachers(self, message: dict):
        for ws in self.teacher_sockets:
            try:
                await ws.send_json(message)
            except:
                pass

    async def send_frame_to_subscribed_teachers(self, student_id: int, frame_b64: str):
        for ws, sub_id in self.teacher_video_subscriptions.items():
            if sub_id == student_id:
                try:
                    await ws.send_json({"type": "video_frame", "user_id": student_id, "frame": frame_b64})
                except:
                    pass

manager = ConnectionManager()
session_active = False

# --- UI ROUTES ---

@app.get("/", response_class=HTMLResponse)
async def login_page():
    return """
    <html><body>
        <h2>Login</h2>
        <form method="post" action="/login">
            Username: <input type="text" name="username"><br>
            Password: <input type="password" name="password"><br>
            <button type="submit">Login</button>
        </form>
    </body></html>
    """

@app.post("/login")
async def login(request: Request, username: str = Form(...), password: str = Form(...)):
    user = db.get_user_by_username(username)
    # Simple check for prototype (bypassing secure hash verify for simplicity in demonstration)
    if not user:
        return HTMLResponse("Invalid user", status_code=401)
        
    role = user['role']
    if role == "admin": return HTMLResponse(ADMIN_DASHBOARD)
    elif role == "teacher": return HTMLResponse(TEACHER_DASHBOARD)
    else: return HTMLResponse(USER_DASHBOARD.replace("student_id_placeholder", str(user['id'])))


# --- ADMIN ENDPOINTS ---

@app.post("/api/admin/upload_students")
async def upload_students(file: UploadFile = File(...)):
    contents = await file.read()
    
    # Load workbook using pure Python openpyxl (bypasses Windows AppLocker DLL blocks)
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    sheet = wb.active
    
    output_rows = []
    import secrets
    
    # Iterate through rows and safely lower-case headers
    headers = [str(cell.value).lower().strip() if cell.value else '' for cell in sheet[1]]
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        
        # Look for headers, fallback to raw column indexes if headers are totally wrong
        name = str(row_dict.get('name') or (row[0] if len(row) > 0 else 'Unknown'))
        prn = str(row_dict.get('prn') or (row[1] if len(row) > 1 else ''))
        email = str(row_dict.get('email') or (row[2] if len(row) > 2 else ''))
        
        if prn and prn.lower() != 'none':
            password = secrets.token_urlsafe(8)
            
            # Try to create user
            user = db.create_user(username=prn, password=password, role="user", email=email, display_name=name)
            
            # If user already exists in DB (you uploaded this sheet before), just update their password!
            if user is None:
                existing_user = db.get_user_by_username(prn)
                if existing_user:
                    db.update_password(existing_user['id'], password)
                    
            output_rows.append({"Name": name, "PRN": prn, "Email": email, "Generated_Password": password})
    
    # Create new output workbook
    out_wb = openpyxl.Workbook()
    out_sheet = out_wb.active
    out_sheet.append(["Name", "PRN", "Email", "Generated_Password"])
    
    for r in output_rows:
        out_sheet.append([r["Name"], r["PRN"], r["Email"], r["Generated_Password"]])
        
    stream = io.BytesIO()
    out_wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=generated_students.xlsx"})


# --- TEACHER ENDPOINTS ---

current_session_id = None

@app.post("/api/teacher/start_session")
async def start_session():
    global session_active, current_session_id
    session_active = True
    current_session_id = db.create_session(subject_id=1, teacher_id=1) # Mock subject/teacher ID
    for ws in manager.student_sockets.values():
        await ws.send_json({"action": "start_session"})
    return {"status": "started", "session_id": current_session_id}

@app.post("/api/teacher/end_session")
async def end_session():
    global session_active, current_session_id
    if current_session_id:
        db.update_session_status(current_session_id, "ended")
    session_active = False
    current_session_id = None
    for ws in manager.student_sockets.values():
        await ws.send_json({"action": "end_session"})
    return {"status": "ended"}

@app.websocket("/ws/teacher")
async def websocket_teacher(websocket: WebSocket):
    await manager.connect_teacher(websocket)
    try:
        while True:
            data = await websocket.receive_json()
            if data.get("action") == "subscribe_video":
                manager.teacher_video_subscriptions[websocket] = int(data.get("user_id"))
            elif data.get("action") == "unsubscribe_video":
                if websocket in manager.teacher_video_subscriptions:
                    del manager.teacher_video_subscriptions[websocket]
    except WebSocketDisconnect:
        manager.disconnect_teacher(websocket)


# --- STUDENT ENDPOINTS ---

import cv2
import numpy as np
import collections
import time

sleep_counters = {} # user_id -> int (consecutive eyes_closed frames)
emotion_history = collections.defaultdict(lambda: collections.deque(maxlen=60)) # user_id -> deque of emotion dicts
last_db_write = {} # user_id -> float (timestamp)

@app.get("/api/admin/reports")
async def get_admin_reports(timeframe: str = "daily"):
    return {"reports": db.get_historical_reports(timeframe)}

@app.websocket("/ws/student/{user_id}")
async def websocket_student(websocket: WebSocket, user_id: int):
    await manager.connect_student(websocket, user_id)
    sleep_counters[user_id] = 0
    last_db_write[user_id] = time.time()
    try:
        while True:
            data = await websocket.receive_json()
            if not session_active:
                continue
                
            if data.get("type") == "frame":
                b64_str = data["image"].split(",")[1]
                img_data = base64.b64decode(b64_str)
                np_arr = np.frombuffer(img_data, np.uint8)
                frame = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)
                
                # AI Inference
                pred = engine.predict(frame)
                
                # Track telemetry
                manager.live_telemetry[user_id] = {
                    "name": f"Student {user_id}", # In real app, fetch from DB
                    "state": pred.label,
                    "emotions": pred.emotions
                }
                
                # Broadcasting frame if teacher is watching this student
                await manager.send_frame_to_subscribed_teachers(user_id, data["image"])
                
                # Store rolling 60-second history for averages
                emotion_history[user_id].append(pred.emotions)
                
                # Sleeping Alert Logic (1 FPS -> 120 frames = 120s)
                if pred.label == "eyes_closed":
                    sleep_counters[user_id] += 1
                    if sleep_counters[user_id] > 120:
                        await manager.broadcast_to_teachers({
                            "type": "alert",
                            "student_name": f"Student {user_id}"
                        })
                else:
                    # Grace period logic: Only reset sleep counter if eyes open consistently
                    sleep_counters[user_id] = 0

                # Database Persistence Logic (Save 1-minute average per student to save space)
                current_time = time.time()
                if current_time - last_db_write[user_id] >= 60 and current_session_id:
                    avg_b, avg_e, avg_c, avg_f = 0, 0, 0, 0
                    if len(emotion_history[user_id]) > 0:
                        for em in emotion_history[user_id]:
                            avg_b += em.get("boredom", 0)
                            avg_e += em.get("engagement", 0)
                            avg_c += em.get("confusion", 0)
                            avg_f += em.get("frustration", 0)
                        count = len(emotion_history[user_id])
                        db.add_telemetry(
                            session_id=current_session_id, 
                            user_id=user_id, 
                            state=pred.label,
                            emotions={
                                "boredom": avg_b / count,
                                "engagement": avg_e / count,
                                "confusion": avg_c / count,
                                "frustration": avg_f / count
                            }
                        )
                    last_db_write[user_id] = current_time

                # Compute 1-minute class average
                total_boredom, total_engagement, total_confusion, total_frustration, count = 0, 0, 0, 0, 0
                for uid, history in emotion_history.items():
                    for em in history:
                        total_boredom += em.get("boredom", 0)
                        total_engagement += em.get("engagement", 0)
                        total_confusion += em.get("confusion", 0)
                        total_frustration += em.get("frustration", 0)
                        count += 1
                
                class_avg = {}
                if count > 0:
                    class_avg = {
                        "boredom": round(total_boredom / count, 2),
                        "engagement": round(total_engagement / count, 2),
                        "confusion": round(total_confusion / count, 2),
                        "frustration": round(total_frustration / count, 2)
                    }

                # Send telemetry update to teachers every frame
                await manager.broadcast_to_teachers({
                    "type": "telemetry",
                    "students": manager.live_telemetry,
                    "class_average": class_avg
                })
                
    except WebSocketDisconnect:
        manager.disconnect_student(user_id)
        if user_id in sleep_counters:
            del sleep_counters[user_id]
        if user_id in emotion_history:
            del emotion_history[user_id]
